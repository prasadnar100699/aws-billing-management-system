from flask import Blueprint, request, render_template, jsonify, send_file
from flask_jwt_extended import jwt_required, get_jwt_identity
from datetime import datetime, timedelta
from sqlalchemy import func, extract
from app import db
from app.models.users import User
from app.models.clients import Client
from app.models.invoices import Invoice
from app.models.payments import Payment
from app.models.services import Service
from app.models.audit_log import AuditLog
import io
import csv
from openpyxl import Workbook

reports_bp = Blueprint('reports', __name__)

@reports_bp.route('/')
@jwt_required()
def reports_dashboard():
    """Reports dashboard."""
    user_id = get_jwt_identity()
    user = User.query.get(user_id)
    
    # Get date range from request (default to current month)
    end_date = datetime.utcnow().date()
    start_date = end_date.replace(day=1)
    
    if request.args.get('start_date'):
        start_date = datetime.strptime(request.args.get('start_date'), '%Y-%m-%d').date()
    if request.args.get('end_date'):
        end_date = datetime.strptime(request.args.get('end_date'), '%Y-%m-%d').date()
    
    # Get comprehensive analytics
    analytics = get_analytics_data(user, start_date, end_date)
    
    return render_template('reports/dashboard.html', 
                         analytics=analytics,
                         start_date=start_date,
                         end_date=end_date)

@reports_bp.route('/revenue')
@jwt_required()
def revenue_report():
    """Revenue report."""
    user_id = get_jwt_identity()
    user = User.query.get(user_id)
    
    # Get filters from request
    start_date = request.args.get('start_date')
    end_date = request.args.get('end_date')
    client_id = request.args.get('client_id', type=int)
    currency = request.args.get('currency', 'ALL')
    
    query = Invoice.query.filter(Invoice.status.in_(['Paid', 'Sent']))
    
    # Apply filters based on user role
    if user.is_client_manager:
        client_ids = [c.id for c in user.managed_clients]
        query = query.filter(Invoice.client_id.in_(client_ids))
    
    if start_date:
        query = query.filter(Invoice.invoice_date >= datetime.strptime(start_date, '%Y-%m-%d').date())
    if end_date:
        query = query.filter(Invoice.invoice_date <= datetime.strptime(end_date, '%Y-%m-%d').date())
    if client_id:
        query = query.filter(Invoice.client_id == client_id)
    if currency != 'ALL':
        query = query.filter(Invoice.currency == currency)
    
    invoices = query.order_by(Invoice.invoice_date.desc()).all()
    
    # Calculate totals
    total_usd = sum(inv.total_amount for inv in invoices if inv.currency == 'USD')
    total_inr = sum(inv.total_amount for inv in invoices if inv.currency == 'INR')
    
    # Monthly breakdown
    monthly_data = {}
    for invoice in invoices:
        month_key = invoice.invoice_date.strftime('%Y-%m')
        if month_key not in monthly_data:
            monthly_data[month_key] = {'usd': 0, 'inr': 0, 'count': 0}
        
        if invoice.currency == 'USD':
            monthly_data[month_key]['usd'] += float(invoice.total_amount)
        else:
            monthly_data[month_key]['inr'] += float(invoice.total_amount)
        monthly_data[month_key]['count'] += 1
    
    # Get clients for filter
    clients_query = Client.query.filter_by(is_active=True)
    if user.is_client_manager:
        clients_query = user.managed_clients.filter_by(is_active=True)
    clients = clients_query.all()
    
    if request.is_json:
        return jsonify({
            'invoices': [inv.to_dict() for inv in invoices],
            'totals': {
                'usd': float(total_usd),
                'inr': float(total_inr),
                'count': len(invoices)
            },
            'monthly_breakdown': monthly_data
        })
    
    return render_template('reports/revenue.html',
                         invoices=invoices,
                         total_usd=total_usd,
                         total_inr=total_inr,
                         monthly_data=monthly_data,
                         clients=clients,
                         filters={
                             'start_date': start_date,
                             'end_date': end_date,
                             'client_id': client_id,
                             'currency': currency
                         })

@reports_bp.route('/clients')
@jwt_required()
def clients_report():
    """Client analysis report."""
    user_id = get_jwt_identity()
    user = User.query.get(user_id)
    
    query = Client.query.filter_by(is_active=True)
    
    # Filter based on user role
    if user.is_client_manager:
        query = user.managed_clients.filter_by(is_active=True)
    
    clients = query.all()
    
    # Calculate client analytics
    client_analytics = []
    for client in clients:
        total_invoices = client.invoices.count()
        total_billed = sum(inv.total_amount for inv in client.invoices if inv.status != 'Draft')
        total_paid = sum(payment.amount for payment in client.payments)
        outstanding = total_billed - total_paid
        
        # Average invoice value
        paid_invoices = [inv for inv in client.invoices if inv.status in ['Paid', 'Sent']]
        avg_invoice = sum(inv.total_amount for inv in paid_invoices) / len(paid_invoices) if paid_invoices else 0
        
        # Last invoice date
        last_invoice = client.invoices.order_by(Invoice.created_at.desc()).first()
        last_invoice_date = last_invoice.invoice_date if last_invoice else None
        
        client_analytics.append({
            'client': client.to_dict(),
            'total_invoices': total_invoices,
            'total_billed': float(total_billed),
            'total_paid': float(total_paid),
            'outstanding': float(outstanding),
            'avg_invoice_value': float(avg_invoice),
            'last_invoice_date': last_invoice_date.isoformat() if last_invoice_date else None
        })
    
    # Sort by total billed (descending)
    client_analytics.sort(key=lambda x: x['total_billed'], reverse=True)
    
    if request.is_json:
        return jsonify({'clients': client_analytics})
    
    return render_template('reports/clients.html', client_analytics=client_analytics)

@reports_bp.route('/overdue')
@jwt_required()
def overdue_report():
    """Overdue invoices report."""
    user_id = get_jwt_identity()
    user = User.query.get(user_id)
    
    query = Invoice.query.filter(
        Invoice.due_date < datetime.utcnow().date(),
        Invoice.status.in_(['Sent', 'Finalized'])
    )
    
    # Filter based on user role
    if user.is_client_manager:
        client_ids = [c.id for c in user.managed_clients]
        query = query.filter(Invoice.client_id.in_(client_ids))
    
    overdue_invoices = query.order_by(Invoice.due_date.asc()).all()
    
    # Group by days overdue
    overdue_groups = {
        '0-30': [],
        '31-60': [],
        '61-90': [],
        '90+': []
    }
    
    for invoice in overdue_invoices:
        days_overdue = invoice.days_overdue
        if days_overdue <= 30:
            overdue_groups['0-30'].append(invoice)
        elif days_overdue <= 60:
            overdue_groups['31-60'].append(invoice)
        elif days_overdue <= 90:
            overdue_groups['61-90'].append(invoice)
        else:
            overdue_groups['90+'].append(invoice)
    
    # Calculate totals
    total_overdue_amount = sum(inv.outstanding_amount for inv in overdue_invoices)
    
    if request.is_json:
        return jsonify({
            'overdue_invoices': [inv.to_dict() for inv in overdue_invoices],
            'overdue_groups': {
                group: [inv.to_dict() for inv in invoices]
                for group, invoices in overdue_groups.items()
            },
            'total_overdue_amount': float(total_overdue_amount)
        })
    
    return render_template('reports/overdue.html',
                         overdue_invoices=overdue_invoices,
                         overdue_groups=overdue_groups,
                         total_overdue_amount=total_overdue_amount)

@reports_bp.route('/services')
@jwt_required()
def services_report():
    """Service usage and revenue report."""
    user_id = get_jwt_identity()
    user = User.query.get(user_id)
    
    # Get service usage statistics
    services = Service.query.filter_by(is_active=True).all()
    service_stats = []
    
    for service in services:
        line_items = service.invoice_line_items
        
        # Filter by user role
        if user.is_client_manager:
            client_ids = [c.id for c in user.managed_clients]
            line_items = line_items.join(Invoice).filter(Invoice.client_id.in_(client_ids))
        
        usage_count = line_items.count()
        total_quantity = sum(item.quantity for item in line_items) or 0
        total_usd_revenue = sum(item.usd_amount for item in line_items if item.usd_amount) or 0
        total_inr_revenue = sum(item.inr_amount for item in line_items if item.inr_amount) or 0
        
        service_stats.append({
            'service': service.to_dict(),
            'usage_count': usage_count,
            'total_quantity': float(total_quantity),
            'total_usd_revenue': float(total_usd_revenue),
            'total_inr_revenue': float(total_inr_revenue)
        })
    
    # Sort by revenue (USD + INR equivalent)
    service_stats.sort(key=lambda x: x['total_usd_revenue'] + (x['total_inr_revenue'] / 80), reverse=True)
    
    if request.is_json:
        return jsonify({'services': service_stats})
    
    return render_template('reports/services.html', service_stats=service_stats)

@reports_bp.route('/audit')
@jwt_required()
def audit_report():
    """Audit log report."""
    user_id = get_jwt_identity()
    user = User.query.get(user_id)
    
    # Only super admin and auditors can view audit logs
    if not (user.is_super_admin or user.is_auditor):
        if request.is_json:
            return jsonify({'error': 'Insufficient permissions'}), 403
        flash('Insufficient permissions', 'error')
        return redirect(url_for('reports.reports_dashboard'))
    
    page = request.args.get('page', 1, type=int)
    action = request.args.get('action', '')
    resource_type = request.args.get('resource_type', '')
    user_filter = request.args.get('user_id', type=int)
    start_date = request.args.get('start_date')
    end_date = request.args.get('end_date')
    
    query = AuditLog.query
    
    # Apply filters
    if action:
        query = query.filter_by(action=action)
    if resource_type:
        query = query.filter_by(resource_type=resource_type)
    if user_filter:
        query = query.filter_by(user_id=user_filter)
    if start_date:
        query = query.filter(AuditLog.timestamp >= datetime.strptime(start_date, '%Y-%m-%d'))
    if end_date:
        query = query.filter(AuditLog.timestamp <= datetime.strptime(end_date + ' 23:59:59', '%Y-%m-%d %H:%M:%S'))
    
    audit_logs = query.order_by(AuditLog.timestamp.desc())\
        .paginate(page=page, per_page=50, error_out=False)
    
    # Get filter options
    actions = db.session.query(AuditLog.action).distinct().all()
    actions = [action[0] for action in actions]
    
    resource_types = db.session.query(AuditLog.resource_type).distinct().all()
    resource_types = [rt[0] for rt in resource_types]
    
    users = User.query.filter_by(is_active=True).all()
    
    if request.is_json:
        return jsonify({
            'audit_logs': [log.to_dict() for log in audit_logs.items],
            'total': audit_logs.total,
            'pages': audit_logs.pages,
            'current_page': page
        })
    
    return render_template('reports/audit.html',
                         audit_logs=audit_logs,
                         actions=actions,
                         resource_types=resource_types,
                         users=users,
                         filters={
                             'action': action,
                             'resource_type': resource_type,
                             'user_id': user_filter,
                             'start_date': start_date,
                             'end_date': end_date
                         })

@reports_bp.route('/export/<report_type>')
@jwt_required()
def export_report(report_type):
    """Export report to CSV/Excel."""
    user_id = get_jwt_identity()
    user = User.query.get(user_id)
    
    format_type = request.args.get('format', 'csv')  # csv or excel
    
    if report_type == 'revenue':
        data = get_revenue_export_data(user)
        filename = f'revenue_report_{datetime.utcnow().strftime("%Y%m%d_%H%M%S")}'
    elif report_type == 'clients':
        data = get_clients_export_data(user)
        filename = f'clients_report_{datetime.utcnow().strftime("%Y%m%d_%H%M%S")}'
    elif report_type == 'overdue':
        data = get_overdue_export_data(user)
        filename = f'overdue_report_{datetime.utcnow().strftime("%Y%m%d_%H%M%S")}'
    else:
        return jsonify({'error': 'Invalid report type'}), 400
    
    if format_type == 'excel':
        return export_to_excel(data, filename)
    else:
        return export_to_csv(data, filename)

def get_analytics_data(user, start_date, end_date):
    """Get comprehensive analytics data."""
    analytics = {}
    
    # Base queries based on user role
    if user.is_client_manager:
        client_ids = [c.id for c in user.managed_clients]
        invoices_query = Invoice.query.filter(Invoice.client_id.in_(client_ids))
        clients_query = user.managed_clients
    else:
        invoices_query = Invoice.query
        clients_query = Client.query
    
    # Date filter
    invoices_query = invoices_query.filter(
        Invoice.invoice_date >= start_date,
        Invoice.invoice_date <= end_date
    )
    
    # Basic counts
    analytics['total_clients'] = clients_query.filter_by(is_active=True).count()
    analytics['total_invoices'] = invoices_query.count()
    analytics['draft_invoices'] = invoices_query.filter_by(status='Draft').count()
    analytics['sent_invoices'] = invoices_query.filter_by(status='Sent').count()
    analytics['paid_invoices'] = invoices_query.filter_by(status='Paid').count()
    
    # Revenue calculations
    revenue_invoices = invoices_query.filter(Invoice.status.in_(['Paid', 'Sent'])).all()
    analytics['total_revenue_usd'] = sum(inv.total_amount for inv in revenue_invoices if inv.currency == 'USD')
    analytics['total_revenue_inr'] = sum(inv.total_amount for inv in revenue_invoices if inv.currency == 'INR')
    
    # Outstanding amounts
    outstanding_invoices = invoices_query.filter_by(status='Sent').all()
    analytics['outstanding_amount'] = sum(inv.outstanding_amount for inv in outstanding_invoices)
    
    # Overdue invoices
    analytics['overdue_invoices'] = invoices_query.filter(
        Invoice.due_date < datetime.utcnow().date(),
        Invoice.status.in_(['Sent', 'Finalized'])
    ).count()
    
    return analytics

def get_revenue_export_data(user):
    """Get revenue data for export."""
    query = Invoice.query.filter(Invoice.status.in_(['Paid', 'Sent']))
    
    if user.is_client_manager:
        client_ids = [c.id for c in user.managed_clients]
        query = query.filter(Invoice.client_id.in_(client_ids))
    
    invoices = query.order_by(Invoice.invoice_date.desc()).all()
    
    data = {
        'headers': ['Invoice Number', 'Client', 'Date', 'Due Date', 'Currency', 'Amount', 'Status', 'Days Overdue'],
        'rows': []
    }
    
    for invoice in invoices:
        data['rows'].append([
            invoice.invoice_number,
            invoice.client.name,
            invoice.invoice_date.strftime('%Y-%m-%d'),
            invoice.due_date.strftime('%Y-%m-%d'),
            invoice.currency,
            float(invoice.total_amount),
            invoice.status,
            invoice.days_overdue if invoice.is_overdue else 0
        ])
    
    return data

def get_clients_export_data(user):
    """Get client data for export."""
    query = Client.query.filter_by(is_active=True)
    
    if user.is_client_manager:
        query = user.managed_clients.filter_by(is_active=True)
    
    clients = query.all()
    
    data = {
        'headers': ['Client Name', 'Email', 'Total Invoices', 'Total Billed', 'Total Paid', 'Outstanding'],
        'rows': []
    }
    
    for client in clients:
        total_billed = sum(inv.total_amount for inv in client.invoices if inv.status != 'Draft')
        total_paid = sum(payment.amount for payment in client.payments)
        outstanding = total_billed - total_paid
        
        data['rows'].append([
            client.name,
            client.email,
            client.invoices.count(),
            float(total_billed),
            float(total_paid),
            float(outstanding)
        ])
    
    return data

def get_overdue_export_data(user):
    """Get overdue invoices data for export."""
    query = Invoice.query.filter(
        Invoice.due_date < datetime.utcnow().date(),
        Invoice.status.in_(['Sent', 'Finalized'])
    )
    
    if user.is_client_manager:
        client_ids = [c.id for c in user.managed_clients]
        query = query.filter(Invoice.client_id.in_(client_ids))
    
    invoices = query.order_by(Invoice.due_date.asc()).all()
    
    data = {
        'headers': ['Invoice Number', 'Client', 'Due Date', 'Amount', 'Days Overdue', 'Outstanding'],
        'rows': []
    }
    
    for invoice in invoices:
        data['rows'].append([
            invoice.invoice_number,
            invoice.client.name,
            invoice.due_date.strftime('%Y-%m-%d'),
            float(invoice.total_amount),
            invoice.days_overdue,
            float(invoice.outstanding_amount)
        ])
    
    return data

def export_to_csv(data, filename):
    """Export data to CSV."""
    output = io.StringIO()
    writer = csv.writer(output)
    
    writer.writerow(data['headers'])
    writer.writerows(data['rows'])
    
    output.seek(0)
    return send_file(
        io.BytesIO(output.getvalue().encode('utf-8')),
        mimetype='text/csv',
        as_attachment=True,
        download_name=f'{filename}.csv'
    )

def export_to_excel(data, filename):
    """Export data to Excel."""
    wb = Workbook()
    ws = wb.active
    ws.title = "Report"
    
    # Add headers
    for col, header in enumerate(data['headers'], 1):
        ws.cell(row=1, column=col, value=header)
    
    # Add data
    for row_idx, row_data in enumerate(data['rows'], 2):
        for col_idx, cell_data in enumerate(row_data, 1):
            ws.cell(row=row_idx, column=col_idx, value=cell_data)
    
    # Save to BytesIO
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    
    return send_file(
        output,
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        as_attachment=True,
        download_name=f'{filename}.xlsx'
    )